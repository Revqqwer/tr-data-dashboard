# -*- coding: utf-8 -*-
"""
Varlik Sinifi (asset-class) kumulatif akis grafigini FON BAZINDA acar ve Excel'e yazar.

Grafik her fonun net akisini, o gunku kompozisyon agirligina gore varlik siniflarina
dagitir (asset_class_history ile ayni mantik). Bu script ayni hesabi fon fon yapip
her fonun:
  - donem kumulatif net akisini,
  - her varlik sinifina yaptigi kumulatif katkiyi,
  - son kompozisyonundaki baskin sinifi (siniflandirma dogrulugu icin)
gosterir.

Kullanim (PythonAnywhere Bash console):
  cd ~/tr-data-dashboard
  python scripts/export_asset_class_funds.py 2025-01-17 2026-07-06
  # opsiyonel 3. arg fund_type: YAT / EMK / BYF
  # cikti: ~/tr-data-dashboard/varlik_sinifi_fonlar.xlsx  (Files sekmesinden indir)
"""
import sys
import os
import bisect
import datetime
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sqlmodel import Session, select
from tefas_backend.database import engine, FundFlow, FundComposition, FundMeta
from tefas_backend.flow_analysis import ASSET_CLASSES


def compute_fund_breakdown(session, start_date, end_date, fund_type=None):
    """asset_class_history ile birebir ayni attribution'i FON BAZINDA dondurur.

    Donen: (per_fund, class_totals)
      per_fund: {code: {"cum_net": float, "contrib": {sinif: float},
                        "kapsanmayan": float,
                        "last_dom": (sinif|None, weight%),}}
      class_totals: {sinif: toplam_kumulatif_katki}
    """
    # --- Akislar: hafif kolonlar, koda gore grupla ---
    fq = select(FundFlow.trade_date, FundFlow.code, FundFlow.net_flow).where(
        FundFlow.trade_date >= start_date,
        FundFlow.trade_date <= end_date,
        FundFlow.net_flow.isnot(None),  # type: ignore
    )
    if fund_type:
        fq = fq.where(FundFlow.fund_type == fund_type.upper())

    flows_by_code = defaultdict(list)  # code -> [(td, net), ...]
    codes = set()
    for td, code, net in session.exec(fq):
        flows_by_code[code].append((td, net or 0.0))
        codes.add(code)
    if not codes:
        return {}, {}

    # --- Kompozisyonlar: sadece gerekli agirlik kolonlari, sinif agirligina indir ---
    comp_cols = []
    for fs in ASSET_CLASSES.values():
        comp_cols += fs
    comp_cols = list(dict.fromkeys(comp_cols))
    cidx = {f: i for i, f in enumerate(comp_cols)}
    lookback = datetime.timedelta(days=45)
    cq = select(
        FundComposition.code, FundComposition.trade_date,
        *[getattr(FundComposition, f) for f in comp_cols],
    ).where(
        FundComposition.trade_date >= start_date - lookback,
        FundComposition.trade_date <= end_date,
        FundComposition.code.in_(list(codes)),  # type: ignore
    )
    tmp = defaultdict(list)  # code -> [(td, {sinif: w}, total_weight)]
    for row in session.exec(cq):
        code = row[0]
        td = row[1]
        vals = row[2:]
        acw = {}
        tw = 0.0
        for ac_name, fields in ASSET_CLASSES.items():
            w = 0.0
            for f in fields:
                v = vals[cidx[f]]
                if v:
                    w += v
            acw[ac_name] = w
            tw += w
        tmp[code].append((td, acw, tw))

    comp_dates = {}
    comp_items = {}
    for code, lst in tmp.items():
        lst.sort(key=lambda x: x[0])
        comp_dates[code] = [x[0] for x in lst]
        comp_items[code] = [(x[1], x[2]) for x in lst]

    def as_of(code, d):
        ds = comp_dates.get(code)
        if not ds:
            return None
        i = bisect.bisect_right(ds, d) - 1
        if i < 0:
            return None
        return comp_items[code][i]

    per_fund = {}
    class_totals = {ac: 0.0 for ac in ASSET_CLASSES}
    for code, flows in flows_by_code.items():
        contrib = {ac: 0.0 for ac in ASSET_CLASSES}
        cum_net = 0.0
        attributed = 0.0
        for td, net in flows:
            cum_net += net
            comp = as_of(code, td)
            if not comp or comp[1] < 1.0:
                continue
            for ac_name, w in comp[0].items():
                if w:
                    part = net * w / 100.0
                    contrib[ac_name] += part
                    attributed += part
        for ac in ASSET_CLASSES:
            class_totals[ac] += contrib[ac]
        # son kompozisyonun baskin sinifi (siniflandirma dogrulugu icin)
        last_dom = (None, 0.0)
        items = comp_items.get(code)
        if items:
            last_acw = items[-1][0]
            if last_acw:
                dom = max(last_acw.items(), key=lambda kv: kv[1])
                last_dom = (dom[0], dom[1])
        per_fund[code] = {
            "cum_net": cum_net,
            "contrib": contrib,
            "kapsanmayan": cum_net - attributed,
            "last_dom": last_dom,
        }
    return per_fund, class_totals


def _write_excel(path, per_fund, class_totals, meta, start_date, end_date, fund_type):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    classes = list(ASSET_CLASSES.keys())
    TL_FMT = '#,##0;(#,##0);-'
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    base_font = Font(name="Arial")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(bottom=thin)

    wb = Workbook()

    # ---- Sheet 1: Fon Bazinda ----
    ws = wb.active
    ws.title = "Fon Bazinda"
    headers = ["Kod", "Fon Adi", "Meta Kategori", "Fon Tipi",
               "Kumulatif Net Akis (TL)", "Baskin Katki Sinifi (donem)",
               "Son Komp. Baskin Sinif", "Son Komp. Agirlik (%)"] + classes + ["Kapsanmayan (TL)"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = []
    for code, d in per_fund.items():
        m = meta.get(code, {})
        dom_class = max(d["contrib"].items(), key=lambda kv: abs(kv[1]))[0] if d["contrib"] else ""
        rows.append((code, m.get("fname", ""), m.get("category", ""), m.get("fund_type", ""),
                     d["cum_net"], dom_class, d["last_dom"][0] or "", round(d["last_dom"][1], 1),
                     d["contrib"], d["kapsanmayan"]))
    rows.sort(key=lambda r: abs(r[4]), reverse=True)

    for r in rows:
        code, fname, cat, ftype, cum_net, dom_class, last_dom, last_w, contrib, kaps = r
        vals = [code, fname, cat, ftype, cum_net, dom_class, last_dom, last_w]
        vals += [contrib[ac] for ac in classes]
        vals.append(kaps)
        ws.append(vals)
    # formatlar
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = base_font
            cell.border = border
        row[4].number_format = TL_FMT            # kumulatif net
        row[7].number_format = '0.0'             # agirlik %
        for i in range(8, 8 + len(classes) + 1): # sinif katkilari + kapsanmayan
            row[i].number_format = TL_FMT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), ws.max_row)
    widths = [8, 34, 20, 8, 20, 24, 22, 16] + [15] * len(classes) + [16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 2: Sinif Ozeti ----
    ws2 = wb.create_sheet("Sinif Ozeti")
    ws2.append(["Varlik Sinifi", "Kumulatif Toplam (TL)", "Katki Yapan Fon Sayisi"])
    for c in range(1, 4):
        cell = ws2.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    counts = {ac: 0 for ac in classes}
    for d in per_fund.values():
        for ac in classes:
            if abs(d["contrib"].get(ac, 0.0)) >= 1.0:
                counts[ac] += 1
    summ = sorted(classes, key=lambda ac: abs(class_totals.get(ac, 0.0)), reverse=True)
    for ac in summ:
        ws2.append([ac, class_totals.get(ac, 0.0), counts[ac]])
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        row[0].font = base_font
        row[1].font = base_font
        row[2].font = base_font
        row[1].number_format = TL_FMT
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 22

    # ---- Sheet 3: Bilgi ----
    ws3 = wb.create_sheet("Bilgi")
    info = [
        ("Donem", "%s -> %s" % (start_date, end_date)),
        ("Fon Tipi filtresi", fund_type or "TUMU"),
        ("Fon sayisi", len(per_fund)),
        ("Not", "Her fonun akisi gunluk kompozisyon agirligina gore siniflara dagitildi."),
        ("Not", "'Kapsanmayan' = kompozisyonu olmayan ya da toplam agirligi %1 alti gunlerin akisi."),
        ("Not", "Sinif katki kolonlarinin toplami = fonun net akisi - Kapsanmayan."),
    ]
    for k, v in info:
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 80

    wb.save(path)


def main():
    args = sys.argv[1:]
    start_date = datetime.date.fromisoformat(args[0]) if len(args) > 0 else datetime.date(2025, 1, 17)
    end_date = datetime.date.fromisoformat(args[1]) if len(args) > 1 else datetime.date(2026, 7, 6)
    fund_type = args[2] if len(args) > 2 else None

    with Session(engine) as s:
        per_fund, class_totals = compute_fund_breakdown(s, start_date, end_date, fund_type)
        meta = {m.code: {"fname": m.fname, "category": m.category, "fund_type": m.fund_type}
                for m in s.exec(select(FundMeta))}

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "varlik_sinifi_fonlar.xlsx")
    _write_excel(out, per_fund, class_totals, meta, start_date, end_date, fund_type)
    print("Fon sayisi:", len(per_fund))
    print("Sinif toplamlari (kumulatif TL):")
    for ac in sorted(class_totals, key=lambda a: abs(class_totals[a]), reverse=True):
        print(f"  {ac:<26} {class_totals[ac]:+,.0f}")
    print("Excel yazildi:", out)


if __name__ == "__main__":
    main()
