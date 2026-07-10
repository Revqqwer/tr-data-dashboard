# -*- coding: utf-8 -*-
"""
Tum fonlarin listesini Excel'e yazar — Istatistikler sayfasi icin fon secmek uzere.

Kolonlar: Kod, Fon Adi, Kategori, Fon Tipi, Son Hisse Agirligi (%), Son Fiyat Tarihi,
          Su An Dahil (Hisse Yogun mu).
Siralama: once su an dahil olanlar, sonra hisse agirligi yuksekten dusuge.

Kullanim (PythonAnywhere Bash console):
  cd ~/tr-data-dashboard && python scripts/export_fund_list.py
  # cikti: ~/tr-data-dashboard/fon_listesi.xlsx  (Files sekmesinden indir)
"""
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sqlmodel import Session, select
from sqlalchemy import func as _func
from tefas_backend.database import engine, FundMeta, FundDaily, FundComposition

CAT_HISSE = "Hisse Senedi Şemsiye Fonu"


def main():
    with Session(engine) as db:
        metas = db.exec(select(FundMeta.code, FundMeta.fname, FundMeta.category, FundMeta.fund_type)).all()

        # Son fiyat tarihi (kod basi)
        last_price = {c: d for c, d in db.exec(
            select(FundDaily.code, _func.max(FundDaily.trade_date)).group_by(FundDaily.code)
        ).all()}

        # Son hisse (hs) agirligi: (kod, tarih, hs) hepsini cek, kod basi en guncel
        last_hs = {}
        last_hs_date = {}
        for code, td, hs in db.exec(select(FundComposition.code, FundComposition.trade_date, FundComposition.hs)):
            if td is None:
                continue
            if code not in last_hs_date or td > last_hs_date[code]:
                last_hs_date[code] = td
                last_hs[code] = hs

    rows = []
    for code, fname, category, ftype in metas:
        included = (category == CAT_HISSE)
        rows.append({
            "code": code,
            "fname": fname or "",
            "category": category or "",
            "ftype": ftype or "",
            "hs": last_hs.get(code),
            "last_px": last_price.get(code),
            "included": included,
        })

    # once dahil olanlar, sonra hisse agirligi yuksek
    rows.sort(key=lambda r: (not r["included"], -(r["hs"] or 0.0)))

    _write_excel(rows)
    n_inc = sum(1 for r in rows if r["included"])
    print(f"Toplam fon: {len(rows)} | Su an dahil (Hisse Yogun): {n_inc}")
    print("Excel:", os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "fon_listesi.xlsx"))


def _write_excel(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Fonlar"
    headers = ["Kod", "Fon Adı", "Kategori", "Fon Tipi",
               "Son Hisse Ağırlığı (%)", "Son Fiyat Tarihi", "Şu An Dahil"]
    ws.append(headers)

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    base_font = Font(name="Arial")
    inc_fill = PatternFill("solid", fgColor="E2F0D9")   # dahil olan satir yesilimsi
    thin = Side(style="thin", color="D9D9D9")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        ws.append([
            r["code"], r["fname"], r["category"], r["ftype"],
            round(r["hs"], 1) if r["hs"] is not None else None,
            r["last_px"].isoformat() if r["last_px"] else "",
            "Evet" if r["included"] else "Hayır",
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = base_font
            cell.border = Border(bottom=thin)
        row[4].number_format = '0.0'
        if row[6].value == "Evet":
            for cell in row:
                cell.fill = inc_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), ws.max_row)
    widths = [9, 40, 30, 9, 20, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "fon_listesi.xlsx")
    wb.save(out)


if __name__ == "__main__":
    main()
